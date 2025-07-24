#!/usr/bin/env python3
"""
MongoDB Hash Comparison Script

This script compares hash values between source and destination MongoDB clusters
using the dbHash command and generates an Excel report highlighting mismatches.

Author: Generated for MongoDB cluster comparison
Date: 2025-07-24
"""

import pymongo
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils.dataframe import dataframe_to_rows
import argparse
import logging
from datetime import datetime
import sys
import os
from typing import Dict, List, Tuple, Any
from dotenv import load_dotenv

# Load environment variables from .env file
load_dotenv()

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('mongodb_hash_compare.log'),
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger(__name__)

class MongoDBHashComparer:
    """Class to handle MongoDB hash comparison between source and destination clusters"""
    
    def __init__(self, source_uri: str, dest_uri: str):
        """
        Initialize the MongoDB hash comparer
        
        Args:
            source_uri: MongoDB connection string for source cluster
            dest_uri: MongoDB connection string for destination cluster
        """
        self.source_uri = source_uri
        self.dest_uri = dest_uri
        self.source_client = None
        self.dest_client = None
        
    def connect_to_clusters(self) -> bool:
        """
        Establish connections to both MongoDB clusters
        
        Returns:
            bool: True if both connections successful, False otherwise
        """
        try:
            logger.info("Connecting to source cluster...")
            self.source_client = pymongo.MongoClient(self.source_uri, serverSelectionTimeoutMS=30000)
            # Test connection
            self.source_client.admin.command('ping')
            logger.info("Source cluster connection successful")
            
            logger.info("Connecting to destination cluster...")
            self.dest_client = pymongo.MongoClient(self.dest_uri, serverSelectionTimeoutMS=30000)
            # Test connection
            self.dest_client.admin.command('ping')
            logger.info("Destination cluster connection successful")
            
            return True
            
        except Exception as e:
            logger.error(f"Failed to connect to clusters: {str(e)}")
            return False
    
    def get_non_system_databases(self, client: pymongo.MongoClient) -> List[str]:
        """
        Get list of non-system databases from MongoDB cluster
        
        Args:
            client: MongoDB client instance
            
        Returns:
            List of database names excluding system databases
        """
        try:
            all_dbs = client.list_database_names()
            # Exclude system databases
            system_dbs = {'admin', 'local', 'config'}
            non_system_dbs = [db for db in all_dbs if db not in system_dbs]
            logger.info(f"Found {len(non_system_dbs)} non-system databases: {non_system_dbs}")
            return non_system_dbs
            
        except Exception as e:
            logger.error(f"Failed to get database list: {str(e)}")
            return []
    
    def run_db_hash(self, client: pymongo.MongoClient, database: str) -> Dict[str, Any]:
        """
        Run dbHash command on a specific database
        
        Args:
            client: MongoDB client instance
            database: Database name
            
        Returns:
            Dictionary containing hash results or empty dict on error
        """
        try:
            logger.info(f"Running dbHash on database: {database}")
            db = client[database]
            result = db.command("dbHash")
            
            # Extract relevant information
            hash_info = {
                'database': database,
                'host': result.get('host', 'unknown'),
                'collections': result.get('collections', {}),
                'md5': result.get('md5', ''),
                'timeMillis': result.get('timeMillis', 0),
                'timestamp': datetime.now().isoformat()
            }
            
            logger.info(f"dbHash completed for {database} in {hash_info['timeMillis']}ms")
            logger.info(f"Found {len(hash_info['collections'])} collections in {database}")
            
            return hash_info
            
        except Exception as e:
            logger.error(f"Failed to run dbHash on database {database}: {str(e)}")
            return {}
    
    def collect_all_hashes(self) -> Tuple[Dict[str, Dict], Dict[str, Dict]]:
        """
        Collect hash information from both source and destination clusters
        
        Returns:
            Tuple containing (source_hashes, dest_hashes) dictionaries
        """
        source_hashes = {}
        dest_hashes = {}
        
        # Get databases from source cluster
        source_dbs = self.get_non_system_databases(self.source_client)
        
        # Get databases from destination cluster  
        dest_dbs = self.get_non_system_databases(self.dest_client)
        
        # Collect hashes from source cluster
        logger.info("Collecting hashes from source cluster...")
        for db_name in source_dbs:
            hash_result = self.run_db_hash(self.source_client, db_name)
            if hash_result:
                source_hashes[db_name] = hash_result
        
        # Collect hashes from destination cluster
        logger.info("Collecting hashes from destination cluster...")
        for db_name in dest_dbs:
            hash_result = self.run_db_hash(self.dest_client, db_name)
            if hash_result:
                dest_hashes[db_name] = hash_result
        
        logger.info(f"Collected hashes for {len(source_hashes)} source databases")
        logger.info(f"Collected hashes for {len(dest_hashes)} destination databases")
        
        return source_hashes, dest_hashes
    
    def prepare_comparison_data(self, source_hashes: Dict, dest_hashes: Dict) -> List[Dict]:
        """
        Prepare data for Excel export by comparing source and destination hashes
        
        Args:
            source_hashes: Hash data from source cluster
            dest_hashes: Hash data from destination cluster
            
        Returns:
            List of dictionaries containing comparison data
        """
        comparison_data = []
        
        # Get all unique database names
        all_databases = set(source_hashes.keys()) | set(dest_hashes.keys())
        
        for db_name in sorted(all_databases):
            source_info = source_hashes.get(db_name, {})
            dest_info = dest_hashes.get(db_name, {})
            
            # Database level comparison
            db_row = {
                'Type': 'Database',
                'Database': db_name,
                'Collection': '',
                'Source_Hash': source_info.get('md5', 'MISSING'),
                'Destination_Hash': dest_info.get('md5', 'MISSING'),
                'Match': 'MISSING DB' if not source_info or not dest_info else 
                        ('MATCH' if source_info.get('md5') == dest_info.get('md5') else 'MISMATCH'),
                'Source_Host': source_info.get('host', 'N/A'),
                'Dest_Host': dest_info.get('host', 'N/A'),
                'Source_Time_ms': source_info.get('timeMillis', 0),
                'Dest_Time_ms': dest_info.get('timeMillis', 0)
            }
            comparison_data.append(db_row)
            
            # Collection level comparison
            source_collections = source_info.get('collections', {})
            dest_collections = dest_info.get('collections', {})
            all_collections = set(source_collections.keys()) | set(dest_collections.keys())
            
            for coll_name in sorted(all_collections):
                source_hash = source_collections.get(coll_name, 'MISSING')
                dest_hash = dest_collections.get(coll_name, 'MISSING')
                
                coll_row = {
                    'Type': 'Collection',
                    'Database': db_name,
                    'Collection': coll_name,
                    'Source_Hash': source_hash,
                    'Destination_Hash': dest_hash,
                    'Match': 'MISSING COLLECTION' if source_hash == 'MISSING' or dest_hash == 'MISSING' else
                            ('MATCH' if source_hash == dest_hash else 'MISMATCH'),
                    'Source_Host': source_info.get('host', 'N/A'),
                    'Dest_Host': dest_info.get('host', 'N/A'),
                    'Source_Time_ms': '',
                    'Dest_Time_ms': ''
                }
                comparison_data.append(coll_row)
        
        return comparison_data
    
    def create_excel_report(self, comparison_data: List[Dict], output_file: str):
        """
        Create Excel report with hash comparison results and highlight mismatches
        
        Args:
            comparison_data: List of comparison data dictionaries
            output_file: Output Excel file path
        """
        try:
            # Create DataFrame
            df = pd.DataFrame(comparison_data)
            
            # Create workbook and worksheet
            wb = Workbook()
            ws = wb.active
            ws.title = "Hash Comparison"
            
            # Define styles
            mismatch_fill = PatternFill(start_color='FFE6E6', end_color='FFE6E6', fill_type='solid')  # Light red
            missing_fill = PatternFill(start_color='FFF0E6', end_color='FFF0E6', fill_type='solid')   # Light orange
            match_fill = PatternFill(start_color='E6F7E6', end_color='E6F7E6', fill_type='solid')     # Light green
            header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')   # Light blue
            header_font = Font(bold=True)
            
            # Add data to worksheet
            for r in dataframe_to_rows(df, index=False, header=True):
                ws.append(r)
            
            # Style header row
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
            
            # Apply conditional formatting
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=len(df) + 1), 2):
                match_status = ws[f'F{row_idx}'].value  # Match column
                
                if 'MISMATCH' in str(match_status):
                    for cell in row:
                        cell.fill = mismatch_fill
                elif 'MISSING' in str(match_status):
                    for cell in row:
                        cell.fill = missing_fill
                elif 'MATCH' == str(match_status):
                    for cell in row:
                        cell.fill = match_fill
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Add summary sheet
            summary_ws = wb.create_sheet("Summary")
            
            # Calculate summary statistics
            total_databases = len([row for row in comparison_data if row['Type'] == 'Database'])
            total_collections = len([row for row in comparison_data if row['Type'] == 'Collection'])
            db_mismatches = len([row for row in comparison_data if row['Type'] == 'Database' and 'MISMATCH' in row['Match']])
            coll_mismatches = len([row for row in comparison_data if row['Type'] == 'Collection' and 'MISMATCH' in row['Match']])
            missing_dbs = len([row for row in comparison_data if row['Type'] == 'Database' and 'MISSING' in row['Match']])
            missing_colls = len([row for row in comparison_data if row['Type'] == 'Collection' and 'MISSING' in row['Match']])
            
            summary_data = [
                ['MongoDB Hash Comparison Summary', ''],
                ['Generated on', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
                ['', ''],
                ['Total Databases Compared', total_databases],
                ['Database Hash Mismatches', db_mismatches],
                ['Missing Databases', missing_dbs],
                ['', ''],
                ['Total Collections Compared', total_collections],
                ['Collection Hash Mismatches', coll_mismatches],
                ['Missing Collections', missing_colls],
                ['', ''],
                ['Overall Status', 'PASS' if (db_mismatches + coll_mismatches + missing_dbs + missing_colls) == 0 else 'FAIL']
            ]
            
            for row_data in summary_data:
                summary_ws.append(row_data)
            
            # Style summary sheet
            summary_ws['A1'].font = Font(bold=True, size=14)
            for row in summary_ws.iter_rows(min_row=1, max_row=len(summary_data)):
                if row[0].value and ':' not in str(row[0].value) and row[0].value != '':
                    row[0].font = Font(bold=True)
            
            # Auto-adjust summary column widths
            for column in summary_ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = max_length + 2
                summary_ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save workbook
            wb.save(output_file)
            logger.info(f"Excel report saved to: {output_file}")
            
            # Print summary to console
            logger.info("=== COMPARISON SUMMARY ===")
            logger.info(f"Total Databases: {total_databases}")
            logger.info(f"Database Mismatches: {db_mismatches}")
            logger.info(f"Missing Databases: {missing_dbs}")
            logger.info(f"Total Collections: {total_collections}")
            logger.info(f"Collection Mismatches: {coll_mismatches}")
            logger.info(f"Missing Collections: {missing_colls}")
            logger.info(f"Overall Status: {'PASS' if (db_mismatches + coll_mismatches + missing_dbs + missing_colls) == 0 else 'FAIL'}")
            
        except Exception as e:
            logger.error(f"Failed to create Excel report: {str(e)}")
            raise
    
    def run_comparison(self, output_file: str = None) -> bool:
        """
        Run the complete hash comparison process
        
        Args:
            output_file: Optional output Excel file path
            
        Returns:
            bool: True if comparison completed successfully
        """
        try:
            if not output_file:
                output_file = f"mongodb_hash_comparison_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            
            # Connect to clusters
            if not self.connect_to_clusters():
                return False
            
            # Collect hashes from both clusters
            source_hashes, dest_hashes = self.collect_all_hashes()
            
            if not source_hashes and not dest_hashes:
                logger.error("No hash data collected from either cluster")
                return False
            
            # Prepare comparison data
            comparison_data = self.prepare_comparison_data(source_hashes, dest_hashes)
            
            # Create Excel report
            self.create_excel_report(comparison_data, output_file)
            
            return True
            
        except Exception as e:
            logger.error(f"Hash comparison failed: {str(e)}")
            return False
        
        finally:
            # Close connections
            if self.source_client:
                self.source_client.close()
            if self.dest_client:
                self.dest_client.close()

def main():
    """Main function to handle environment variables and run the comparison"""
    # Get configuration from environment variables
    source_uri = os.getenv('SOURCE_MONGODB_URI')
    dest_uri = os.getenv('DEST_MONGODB_URI')
    output_file = os.getenv('OUTPUT_FILE')
    verbose = os.getenv('VERBOSE', 'false').lower() == 'true'
    
    # Validate required environment variables
    if not source_uri:
        logger.error("SOURCE_MONGODB_URI environment variable is required")
        logger.error("Please set it in your .env file or environment")
        sys.exit(1)
    
    if not dest_uri:
        logger.error("DEST_MONGODB_URI environment variable is required")
        logger.error("Please set it in your .env file or environment")
        sys.exit(1)
    
    # Optional: Allow command line overrides
    parser = argparse.ArgumentParser(description='Compare MongoDB cluster hashes using dbHash command')
    parser.add_argument('--source', help='Source MongoDB connection string (overrides env var)')
    parser.add_argument('--destination', help='Destination MongoDB connection string (overrides env var)')
    parser.add_argument('--output', help='Output Excel file path (overrides env var)')
    parser.add_argument('--verbose', '-v', action='store_true', help='Enable verbose logging')
    
    args = parser.parse_args()
    
    # Override environment variables with command line arguments if provided
    if args.source:
        source_uri = args.source
    if args.destination:
        dest_uri = args.destination
    if args.output:
        output_file = args.output
    if args.verbose:
        verbose = True
    
    if verbose:
        logging.getLogger().setLevel(logging.DEBUG)
    
    logger.info("Starting MongoDB Hash Comparison")
    logger.info(f"Source cluster: {source_uri.split('@')[-1] if '@' in source_uri else source_uri}")
    logger.info(f"Destination cluster: {dest_uri.split('@')[-1] if '@' in dest_uri else dest_uri}")
    
    # Create comparer instance
    comparer = MongoDBHashComparer(source_uri, dest_uri)
    
    # Run comparison
    success = comparer.run_comparison(output_file)
    
    if success:
        logger.info("Hash comparison completed successfully!")
        sys.exit(0)
    else:
        logger.error("Hash comparison failed!")
        sys.exit(1)

if __name__ == "__main__":
    main()
